import openpyxl

# for creating a new workbook
#wb = openpyxl.Workbook()

# load an existing workbook
wb = openpyxl.load_workbook("transactions.xlsx")
print("sheet name: ", wb.sheetnames)

sheet = wb['Sheet1']

# to create a new sheet # we can optionally pass index to position this sheet
#wb.create_sheet("Sheet2", 0)
# to remove sheet
# wb.remove_sheet('Sheet2')

print("Rows = ", sheet.max_row)
print("Columns = ", sheet.max_column)

# we can pass coordinates of cell to access cell in sheet
cell1 = sheet['a1']
print("cell a1: ", cell1.value)

# using square bracket we can also access range of cells
column = sheet['a']  # gives us all the cells in col A
print("column A: ", column)  # this returns a tuple of cell objs

cells = sheet["a:c"]  # gives us all cells from col A to col C
# returns tuple of tuples- each tuple representing each column
print("cells from col A to col C: ", cells)

# similarly we can access cell in rows
row = sheet[1]  # cells in row 1
print("row1 = ", row)

cells_inrow = sheet[1:3]  # cells from row1 to row3
print("cells from row1 to row3: ", cells_inrow)

other_cells = sheet["a1:c3"]  # gives us cells from a1 to c3
print("cells from a1 to c3: ", other_cells)

# another way to access cell
cell2 = sheet.cell(row=1, column=1)
# this method is usefull when we are iterating over row n column and
# want to access cells dynamically, see below

print("cell(row=1,column=1) : ", cell2.value)
cell2.value = 't_id'
print("value of cell above changed", cell2.value)

print("cell2.row = ", cell2.row)
print("cell2.column = ", cell2.column)
print("cell2.coordinate = ", cell2.coordinate)

# since in excel rows start from 1, so in range function we give starting index as 1.
print("accessing cells dynamically: ")

for col in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        print(sheet.cell(row, col).value)


# inserting
# to insert a row at the end of the sheet
sheet.append([1, 2, 3])
print("no. of rows after append 1 row: ", sheet.max_row)


print("accessing cells dynamically: ")

for col in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        print(sheet.cell(row, col).value)

# to insert row at a given index
sheet.insert_rows(1)
print("inserting empty row at 1")
# to insert row at a given index
sheet.insert_cols(1)
print("inserting empty col at 1")

print("accessing cells dynamically: ")

for col in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        print(sheet.cell(row, col).value)

sheet.delete_rows(1)
print("deleting row at 1")

sheet.delete_cols(1)
print("deleting col at 1")

print("accessing cells dynamically: ")

for col in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        print(sheet.cell(row, col).value)

wb.save("transactions2.xlsx")
print("Saving..")

# command query separation principle - our methods/ functions should either be
# commands(changes the state of a system) or
# query(access system, doesnt change the state)
# we should not do both things simultaneously

# ex if i print all cell in col 1, moving beyond the data present
# in this case till 9th row, when we only have 4 rows
# after print this result, if we append a row, it would be appended in the 10th row, why?
# because after we iterated over our sheet, the position/index at the end of loop is at the 9th row
# so the row got appended at 10th position
print("col 1 till row 10: \n")
for row in range(1, 10):
    cellp = sheet.cell(row, 1)
    print(cellp.value)
# here since the cell do not exist after 5th row, this code would magically create extra cell.
# this is violation of the above principle

sheet.append([10, 20, 30])
print("no. of rows after append 1 row: ", sheet.max_row)

# so this should be strictly avoided
wb.save("transactions2_error.xlsx")
print("Saving..")
